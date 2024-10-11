from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from apps.logIn.views import group_required
from django.http import JsonResponse
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import Group
from apps.users.models import CustomUser
from apps.activityLog.models import ActivityLog
from django.core.paginator import Paginator
import json
from django.apps import apps
import os
from django.conf import settings
from PIL import Image
from io import BytesIO
from django.core.files.base import ContentFile
from apps.activityLog.utils import log_activity
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from django.db.models import Q


group_name = {
    'Consultor': 'consultants', 
    'Administrador': 'administrators', 
    'Tecnico': 'technicians'
}

status_mapping = {
        'Activo': 'active',
        'Bloqueado': 'blocked',
}

def get_group_by_name(group_name):
    """Obtiene un grupo por su nombre, o None si no existe."""
    try:
        return Group.objects.get(name=group_name)
    except Group.DoesNotExist:
        return None
    

@login_required
@group_required(['administrators'], redirect_url='expired_session')
def view_users(request):
    user_list = CustomUser.objects.exclude(id=request.user.id).order_by('-id')
    search_query = request.GET.get('search', '')

    if search_query:
        user_list = user_list.filter(
        Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query)
    )

    paginator = Paginator(user_list, 13)  # Ajusta el número de usuarios por página
    page_number = request.GET.get('page')
    users = paginator.get_page(page_number)

    # Verifica si es una solicitud AJAX a través del encabezado HTTP
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('partials/_user_table_body.html', {'users': users}, request=request)
        return HttpResponse(html)

    context = {'users': users}
    return render(request, 'view_users.html', context)

@login_required
@group_required(['administrators'], redirect_url='expired_session')
def edit_user(request, id):
    user = get_object_or_404(CustomUser, pk=id)
    activity = ActivityLog.objects.filter(user_id = id).order_by('-timestamp')
    context = {'user': user, 
               'activity': activity}
    return render(request, 'user_edit.html', context)


def validate_user_data(data, is_update=False):
    required_fields = ['names', 'lastName', 'email', 'jobName', 'username', 'status', 'group']

    if not is_update:
        required_fields.extend(['password', 'password_validation'])

    for field in required_fields:
        if not data.get(field):
            return {'success': False, 'error': f'El campo {field} es obligatorio.'}

    if not is_update and data['password'] != data['password_validation']:
        return {'success': False, 'error': 'Las contraseñas no coinciden.'}

    if CustomUser.objects.filter(username=data['username']).exists():
        return {'success': False, 'error': 'El usuario ya existe.'}

    if CustomUser.objects.filter(email=data['email']).exists():
        return {'success': False, 'error': 'El correo ya está registrado.'}

    group_mapping = {
        'Administrador': 'administrators',
        'Consultor': 'consultants',
        'Tecnico': 'technicians',
    }
    if group_mapping.get(data['group'], None) is None:
        return {'success': False, 'error': 'Grupo no válido.'}

    status_mapping = {
        'Activo': 'active',
        'Bloqueado': 'blocked',
    }
    if status_mapping.get(data['status'], None) is None:
        return {'success': False, 'error': 'Estado no válido.'}

    return {'success': True}


@login_required
@group_required(['administrators'], redirect_url='expired_session')
def register_user(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        validation_result = validate_user_data(data)
        if not validation_result['success']:
            return JsonResponse(validation_result)

        try:
            group = get_group_by_name(group_name[data['group']])
            if not group:
                return JsonResponse({'success': False, 'error': 'Grupo no válido.'})

            new_user = CustomUser(
                username=data['username'],
                email=data['email'],
                first_name=data['names'].title(),
                last_name=data['lastName'].title(),
                position=data['jobName'],
                status=status_mapping[data['status']] ,
            )
            new_user.set_password(data['password'])
            new_user.save()
            new_user.groups.add(group)
            

            log_activity(
                user=request.user.id,                       
                action='CREATE',                 
                title='Registro de Usuario',      
                description=f'El usuario registro a {data["names"].title()} en el sistema.',  
                link=f'/edit_user/{new_user.id}',      
                category='USER_PROFILE'          
            )
            messages.success(request, 'Usuario registrado exitosamente.')
            return JsonResponse({'success': True})
        except Exception as e:
            messages.error(request, f'Error al registrar el usuario: {str(e)}')
            return JsonResponse({'success': False, 'error': 'Error interno del servidor.'}, status=500)

    return JsonResponse({'success': False, 'error': 'Método no permitido.'})


@login_required
@group_required(['administrators'], redirect_url='expired_session')
def update_personal_data(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        data = json.loads(request.body)
        errors = {}

        if data['names'] and data['lastName'] and data['email'] and data['email']:

            if 'names' in data and data['names'].strip():
                user.first_name = data['names'].strip().title()

            if 'lastName' in data and data['lastName'].strip():
                user.last_name = data['lastName'].strip().title()

            if 'email' in data and data['email'].strip():
                new_email = data['email'].strip().lower()
                if CustomUser.objects.filter(email=new_email).exclude(id=user.id).exists():
                    return JsonResponse({'success': False, 'error': {'email': 'El correo ya está registrado por otro usuario.'}})

                user.email = new_email

            if 'jobName' in data and data['jobName'].strip():
                user.position = data['jobName'].strip()

            try:
                user.save()
                log_activity(
                    user=request.user.id,                       
                    action='EDIT',                 
                    title='Edito usuario',      
                    description=f'El usuario edito la  informacion personal de {user.first_name}.',  
                    link=f'/edit_user/{user.id}',      
                    category='USER_PROFILE'          
                )
                messages.success(request, 'Datos personales actualizados correctamente.')
                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'error': 'Error interno del servidor.'}, status=500)
        else:
            return JsonResponse({'success': False, 'error': 'Todos los campos son requeridos.'}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
@group_required(['administrators'], redirect_url='expired_session')
def update_login_data(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        data = json.loads(request.body)

        if 'username' in data and data['username']:
            new_username = data['username'].strip()
            if CustomUser.objects.filter(username=new_username).exclude(id=user.id).exists():
                return JsonResponse({'success': False, 'error': 'El usuario ya existe.'})

            user.username = new_username
        else:
            return JsonResponse({'success': False, 'error': 'El usuario es obligatorio.'})

        if 'status' in data:
            
            if data['status'] != 'Estado':  # Si no es 'Estado', significa que hubo un cambio
                if data['status'] == 'Activo':
                    if user.status == 'blocked':
                        log_activity(
                            user=user.id,                       
                            action='LOGIN',                 
                            title='Cuenta desbloqueada',      
                            description=f'La cuenta del usuario ha pasado a estado activo.',  
                            link=f'/edit_user/{user.id}',      
                            category='USER_PROFILE'          
                        )
                    user.status = 'active'
                    user.login_attempts = 0
                elif data['status'] == 'Bloqueado':
                    log_activity(
                        user=user.id,                       
                        action='LOCKOUT',                 
                        title='Cuenta bloqueada',      
                        description=f'La cuenta del del usuario ha sido bloqueada por administrador.',  
                        link=f'/edit_user/{user.id}',      
                        category='USER_PROFILE'          
                    )
                    user.status = 'blocked'
                else:
                    return JsonResponse({'success': False, 'error': 'Error en el estado'})  # Error si no es ni 'Activo' ni 'Bloquedo'
            else:
                # Si data['status'] es 'Estado', simplemente no se hace ningún cambio
                pass

        if 'role' in data:
            group_name = {'Consultor': 'consultants', 'Administrador': 'administrators', 'Tecnico': 'technicians'}.get(data['role'])
            if group_name:
                user.groups.clear()
                group = get_group_by_name(group_name)
                if group:
                    user.groups.add(group)
                else:
                    return JsonResponse({'success': False, 'error': f'El grupo "{group_name}" no existe.'})

        if 'password' in data and data['password'].strip():
            if data['password'] == data.get('password_validation', ''):
                user.set_password(data['password'])
            else:
                return JsonResponse({'success': False, 'error': 'Las contraseñas no coinciden.'})

        user.save()
        log_activity(
            user=request.user.id,                       
            action='EDIT',                 
            title='Edito usuario',      
            description=f'El usuario edito la  informacion de  inicio de {user.first_name}.',  
            link=f'/edit_user/{user.id}',      
            category='USER_PROFILE'          
        )
        messages.success(request, 'Datos de inicio de sesión actualizados correctamente.')
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

def compress_and_convert_to_webp(image_file):
    # Abrir la imagen con Pillow
    img = Image.open(image_file)

    # Convertir la imagen a RGB si es necesario (porque WebP no soporta modo RGBA)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")

    # Comprimir la imagen y reducir su tamaño
    img_io = BytesIO()
    img.save(img_io, format='WEBP', quality=40)  # Cambiar la calidad según sea necesario

    # Crear un archivo Django de la imagen comprimida
    img_content = ContentFile(img_io.getvalue(), name=f"{image_file.name.split('.')[0]}.webp")
    return img_content

@login_required
@group_required(['administrators'], redirect_url='expired_session')
def update_image(request):
    if request.method == 'POST':
        app_name = request.POST.get('app_name')
        table = request.POST.get('table')
        field = request.POST.get('field')
        record_id = request.POST.get('userId')
        new_image = request.FILES.get('image')

        print(app_name, table, field, record_id, new_image)

        if not (app_name and table and field and record_id and new_image):
            return messages.error(request, 'Faltan campos requeridos')

        try:
            record_id = int(record_id)
            Model = apps.get_model(app_name, table)
            instance = Model.objects.get(id=record_id)

            if not hasattr(instance, field):
                return messages.error(request, f'El campo "{field}" no existe en el modelo "{table}".')
            

            current_image_path = getattr(instance, field)

            if current_image_path and current_image_path.url != settings.MEDIA_URL + 'default/default_user.jpg': # Eliminar la imagen anterior si no es la imagen por defecto
                current_image_full_path = os.path.join(settings.MEDIA_ROOT, current_image_path.name)
                if os.path.isfile(current_image_full_path):
                    os.remove(current_image_full_path)

            new_image_name = f'imageProfile_{record_id}.webp'

            compressed_image = compress_and_convert_to_webp(new_image)

            image_file = ContentFile(compressed_image.read(), name=new_image_name)

            setattr(instance, field, image_file)  # Guardar el archivo en el campo
            instance.save()

            messages.success(request, 'Foto de perfil actualizada correctamente')
        except ValueError:
            messages.error(request, 'ID de registro no válido')
        except Model.DoesNotExist:
            messages.error(request, 'El registro no existe')
        except LookupError:
            messages.error(request, f'La app "{app_name}" no tiene un modelo "{table}".')
        except Exception as e:
            messages.error(request, str(e))
    else:
        messages.error(request, 'Método no permitido')

@login_required
@group_required(['administrators'], redirect_url='expired_session')
def delete_user(request, user_id):
    if request.method == 'POST':
        try:
            user_instance = get_object_or_404(CustomUser, id=user_id)

            # Verificar si la imagen no es la por defecto
            if user_instance.profile_picture and user_instance.profile_picture.url != settings.MEDIA_URL + 'default/default_user.jpg':
                # Obtener la ruta completa de la imagen
                image_path = user_instance.profile_picture.path
                # Eliminar la imagen del sistema de archivos
                if os.path.exists(image_path):
                    os.remove(image_path)

            # Eliminar el usuario
            user_instance.delete()
            messages.success(request, 'Usuario eliminado correctamente.')
            log_activity(
                user=request.user.id,                       
                action='DELETE',                 
                title='Elimino usuario',      
                description=f'El usuario elimino el perfil de {user_instance.first_name}.',  
                category='USER_PROFILE'          
            )
            return HttpResponse(status=200)  # Respuesta exitosa
        except Exception as e:
            messages.error(request, str(e))
            return HttpResponse(status=500)  # Error interno en el servidor
    else:
        messages.error(request, 'Método no permitido.')
        return HttpResponse(status=405)  # Método no permitido

@login_required
@group_required(['administrators'], redirect_url='expired_session')
def print_users(request):
    users = CustomUser.objects.all() 
    current_time = datetime.now() 

    context = {
        'users': users,
        'current_time': current_time,  
    }
    
    return render(request, 'prin_table_users.html', context)

@login_required
@group_required(['administrators'], redirect_url='expired_session')
def export_users_xlsx(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Usuarios Registrados'

    headers = ['ID', 'Nombre', 'Correo', 'Usuario', 'Fecha de creación', 'Último inicio de sesión', 'Cargo', 'Estado']

    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    users = CustomUser.objects.all()
    for row_num, user in enumerate(users, start=2): 
        sheet.cell(row=row_num, column=1, value=user.id)
        sheet.cell(row=row_num, column=2, value=f"{user.first_name} {user.last_name}")
        sheet.cell(row=row_num, column=3, value=user.email)
        sheet.cell(row=row_num, column=4, value=user.username)
        sheet.cell(row=row_num, column=5, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row=row_num, column=6, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'N/A')
        sheet.cell(row=row_num, column=7, value=user.position)
        sheet.cell(row=row_num, column=8, value=user.status)

    # Ajustar el ancho de las columnas automáticamente
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in sheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo Excel en un objeto BytesIO
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios_registrados.xlsx'
    workbook.save(response)

    return response